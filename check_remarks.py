#!/usr/bin/env python3
"""检查招生计划 Excel 中的备注，并将结果写入原工作簿的副本。

运行方式：
    python check_remarks.py input.xlsx

依赖：openpyxl
"""

from __future__ import annotations

import argparse
import math
import re
import sys
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter


# 可维护规则：后期新增错字时，只需在这里补充“错误内容: 正确内容”。
TYPO_MAP: dict[str, str] = {
    "详贝院校招生章程": "详见院校招生章程",
    "详贝学校招生章程": "详见学校招生章程",
    "祥见院校招生章程": "详见院校招生章程",
    "祥见学校招生章程": "详见学校招生章程",
    "详见院校招生张程": "详见院校招生章程",
    "详见学校招生张程": "详见学校招生章程",
    "详见院校招牛章程": "详见院校招生章程",
    "详见学校招牛章程": "详见学校招生章程",
    "详见院校招生章呈": "详见院校招生章程",
    "详见学校招生章呈": "详见学校招生章程",
    "详贝": "详见",
    "祥见": "详见",
    "张程": "章程",
    "章呈": "章程",
    "招牛": "招生",
    "召生": "招生",
    "只召": "只招",
    "召收": "招收",
    "身休健康": "身体健康",
    "身体建康": "身体健康",
    "身体健庚": "身体健康",
    "体捡": "体检",
    "休检": "体检",
    "只招有专业志原考生": "只招有专业志愿考生",
    "只招有专亚志愿考生": "只招有专业志愿考生",
    "有专业志原": "有专业志愿",
    "专业志原": "专业志愿",
    "专亚志愿": "专业志愿",
    "专此志愿": "专业志愿",
    "志愿考牛": "志愿考生",
    "考牛": "考生",
    "男牛": "男生",
    "女牛": "女生",
    "语仲": "语种",
    "语钟": "语种",
    "曰语": "日语",
    "英浯": "英语",
    "英诘": "英语",
    "不招色肓": "不招色盲",
    "不招色弱色肓": "不招色弱色盲",
    "色育": "色盲",
    "色肓": "色盲",
    "色若": "色弱",
    "色弱色育": "色弱色盲",
    "色弱色肓": "色弱色盲",
    "色盲色若": "色盲色弱",
    "不直报考": "不宜报考",
    "不宣报考": "不宜报考",
    "不官报考": "不宜报考",
    "进人": "进入",
    "人学": "入学",
    "人校": "入校",
    "转人": "转入",
    "编人": "编入",
    "人读": "入读",
    "单色识别不全者慎报": "单色识别不全者慎报",
    "单色识别不金": "单色识别不全",
    "单色识别不仝": "单色识别不全",
    "单色识别丕全": "单色识别不全",
    "中外合作力学": "中外合作办学",
    "校企合作力学": "校企合作办学",
    "合作力学": "合作办学",
    "中外合作办字": "中外合作办学",
    "校企合作办字": "校企合作办学",
    "联合培荞": "联合培养",
    "办学地占": "办学地点",
    "办学地奌": "办学地点",
    "办学地点详见院校章程": "办学地点详见院校招生章程",
    "师范粪": "师范类",
    "帅范类": "师范类",
    "帅范": "师范",
    "非公费帅范": "非公费师范",
    "公费帅范": "公费师范",
    "囯家专项计划": "国家专项计划",
    "国家专顷计划": "国家专项计划",
    "国家专顶计划": "国家专项计划",
    "地方专顷计划": "地方专项计划",
    "地方专顶计划": "地方专项计划",
    "专项计刘": "专项计划",
    "计刘": "计划",
    "项日": "项目",
    "项自": "项目",
    "顷目": "项目",
    "专顷": "专项",
    "专顶": "专项",
    "少数民旅": "少数民族",
    "民旅": "民族",
    "加份": "加分",
    "政第": "政策",
    "执衍": "执行",
    "认同并执厅": "认同并执行",
    "符台": "符合",
    "台格": "合格",
    "成绩台格": "成绩合格",
    "口试成缋": "口试成绩",
    "成缋": "成绩",
    "录収": "录取",
    "录叹": "录取",
    "包含专亚": "包含专业",
    "包含专此": "包含专业",
    "包含专止": "包含专业",
    "学贵": "学费",
    "学弗": "学费",
    "字费": "学费",
    "住宿贵": "住宿费",
    "收费标淮": "收费标准",
    "标淮": "标准",
}

# 低置信 OCR 疑似错字：只标注，不自动修正，避免把真实内容改坏。
SUSPECT_TYPO_MAP: dict[str, str] = {
    "老生": "考生",
    "老试": "考试",
    "报孝": "报考",
    "慎填": "慎报",
    "校冈": "校区",
    "校医": "校区",
    "由请": "申请",
    "甲请": "申请",
    "电请": "申请",
    "攻策": "政策",
    "识刖": "识别",
    "语神": "语种",
}

# 白名单仅保护合法固定表达，不用于删除或改写备注内容。
WHITELIST = [
    "中外合作办学",
    "校企合作",
    "师范类",
    "地方专项计划",
    "国家专项计划",
    "只招英语语种考生",
    "详见院校招生章程",
    "不招色盲",
    "不招色弱",
    "色盲色弱不宜报考",
]

REMARK_COLUMN_NAMES = ("备注", "专业备注", "计划备注", "招生备注")
EXCEL_EXPORT_COLUMN_WIDTH = 12.75
INVALID_REMARKS = {"无", "暂无", "-", "/"}
EMPTY_TEXT_VALUES = {"nan", "none"}
ENGLISH_PUNCTUATION_MAP = str.maketrans(
    {",": "，", ";": "；", ":": "：", "(": "（", ")": "）"}
)

UNCERTAIN_ABNORMAL_CHARS = {"□", "?", "？", "*"}
CLEARLY_CORRUPT_CHARS = {"�"}
INVISIBLE_CHAR_RE = re.compile(
    r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F\u00AD\u200B-\u200F"
    r"\u202A-\u202E\u2060-\u206F\uFEFF]"
)
HORIZONTAL_SPACE_RE = re.compile(
    r"[\t\f\v \u00A0\u1680\u2000-\u200A\u202F\u205F\u3000]+"
)
HAN_SPACE_HAN_RE = re.compile(r"(?<=[\u3400-\u9FFF]) (?=[\u3400-\u9FFF])")
DELIMITED_SPLIT_RE = re.compile(r"([。；;，,\r\n]+)")
REPEATED_PUNCTUATION_RE = re.compile(r"([，；：。！？、])\1+")
PHRASE_CHAR_RE = re.compile(r"^[\u3400-\u9FFFA-Za-z0-9]+$")
BRACKET_CONTENT_RE = re.compile(r"（([^（）]*)）")
EMPTY_BRACKET_RE = re.compile(r"（\s*）")
REDUNDANT_NESTED_BRACKET_RE = re.compile(r"（\s*（([^（）]*)）\s*）")
BRACKET_GROUP_SEPARATOR_RE = re.compile(r"）[、，；]+（")
LEADING_PUNCTUATION_RE = re.compile(r"^[，；：。！？、]+")
TRAILING_SEPARATOR_RE = re.compile(r"[，；：。！？、]+$")
PUNCTUATION_RUN_RE = re.compile(r"[，；：。！？、]{2,}")
HAN_OCR_NOISE_SYMBOL_RE = re.compile(r"(?<=[\u3400-\u9FFF])[%％](?=[\u3400-\u9FFF])")
SHORT_TRAILING_PUNCTUATION_MAX_CHARS = 4
TUITION_RE = re.compile(
    r"学费[^，；。！？、（）\d]{0,12}?(\d+(?:\.\d+)?)\s*(万元|万|元)"
)
HEIGHT_VALUE_RE = re.compile(
    r"(?:身高|身长)[^，；。！？、（）]{0,12}?(\d+(?:\.\d+)?)\s*(cm|CM|厘米|米|m|M)"
)
HEIGHT_WRONG_UNIT_RE = re.compile(
    r"(?:身高|身长)[^，；。！？、（）]{0,12}?(\d+(?:\.\d+)?)\s*(kg|KG|公斤|千克|斤)"
)
WEIGHT_VALUE_RE = re.compile(
    r"体重[^，；。！？、（）]{0,12}?(\d+(?:\.\d+)?)\s*(kg|KG|公斤|千克|斤)"
)
WEIGHT_WRONG_UNIT_RE = re.compile(
    r"体重[^，；。！？、（）]{0,12}?(\d+(?:\.\d+)?)\s*(cm|CM|厘米|米|m|M)"
)

BRACKET_CONTENT_TYPO_MAP: dict[str, str] = {
    "通类": "普通类",
    "普类": "普通类",
}


def _unique(items: list[str]) -> list[str]:
    """按首次出现顺序去重。"""
    return list(dict.fromkeys(item for item in items if item))


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    try:
        if value != value:
            return ""
    except (TypeError, ValueError):
        pass
    return str(value)


def _is_empty_remark(value: Any) -> bool:
    text = _to_text(value).strip()
    return not text or text.lower() in EMPTY_TEXT_VALUES


def _is_invalid_remark(value: Any) -> bool:
    return _to_text(value).strip() in INVALID_REMARKS


def find_remark_column(source: Any) -> str:
    """按约定列名自动寻找备注列，找不到时抛出清晰错误。"""
    columns = source.columns if hasattr(source, "columns") else source
    normalized = {str(column).strip(): str(column) for column in columns if column is not None}
    for candidate in REMARK_COLUMN_NAMES:
        if candidate in normalized:
            return normalized[candidate]
    expected = "、".join(REMARK_COLUMN_NAMES)
    raise ValueError(f"未找到备注列。请确认表头包含以下列名之一：{expected}")


def _describe_abnormal_char(char: str) -> str:
    if INVISIBLE_CHAR_RE.fullmatch(char):
        return f"不可见字符 U+{ord(char):04X}"
    return char


def check_abnormal_chars(text: str) -> tuple[str, list[str], bool]:
    """删除明确乱码；不确定字符只标注，不强制删除。"""
    found: list[str] = []
    cleaned: list[str] = []

    for char in text:
        code_point = ord(char)
        is_invisible = bool(INVISIBLE_CHAR_RE.fullmatch(char))
        is_corrupt = char in CLEARLY_CORRUPT_CHARS
        is_uncertain = char in UNCERTAIN_ABNORMAL_CHARS
        is_private_use = 0xE000 <= code_point <= 0xF8FF

        if is_invisible or is_corrupt or is_uncertain or is_private_use:
            found.append(_describe_abnormal_char(char))

        if not (is_invisible or is_corrupt or is_private_use):
            cleaned.append(char)

    fixed = "".join(cleaned)
    abnormal = _unique(found)
    issues = [f"存在疑似异常字符：{'、'.join(abnormal)}"] if abnormal else []
    return fixed, issues, fixed != text


def check_typos(text: str) -> tuple[str, list[str], bool]:
    current = text
    issues: list[str] = []

    for wrong, correct in sorted(TYPO_MAP.items(), key=lambda item: len(item[0]), reverse=True):
        if wrong == correct or wrong not in current:
            continue
        current = current.replace(wrong, correct)
        issues.append(f"疑似错字：{wrong} → {correct}")

    return current, issues, current != text


def check_suspect_typos(text: str) -> tuple[str, list[str], bool]:
    issues: list[str] = []

    for wrong, correct in sorted(
        SUSPECT_TYPO_MAP.items(), key=lambda item: len(item[0]), reverse=True
    ):
        if wrong in text:
            issues.append(f"疑似 OCR 错字：{wrong} 可能为 {correct}，请人工检查")

    return text, issues, False


def check_ocr_noise_symbols(text: str) -> tuple[str, list[str], bool]:
    current = HAN_OCR_NOISE_SYMBOL_RE.sub("", text)
    issues = ["疑似 OCR 多余符号：%/％"] if current != text else []
    return current, issues, current != text


def _duplicate_key(segment: str) -> str:
    return re.sub(r"[\s\u3000]+", "", segment).strip()


def _collapse_punctuation_run(match: re.Match[str]) -> str:
    run = match.group(0)
    terminal_marks = [char for char in run if char in "。！？"]
    if terminal_marks:
        return terminal_marks[-1]
    if "；" in run:
        return "；"
    if "，" in run:
        return "，"
    return run[0]


def _trailing_meaningful_length(text: str) -> int:
    segments = [segment for segment in re.split(r"[，；：。！？、（）\s]+", text.strip()) if segment]
    if not segments:
        return 0
    return sum(1 for char in segments[-1] if char.isalnum())


def _should_strip_short_trailing_punctuation(text_before_punctuation: str) -> bool:
    length = _trailing_meaningful_length(text_before_punctuation)
    return 0 < length <= SHORT_TRAILING_PUNCTUATION_MAX_CHARS


def _strip_short_punctuation_before_closing_bracket(text: str) -> str:
    def replace(match: re.Match[str]) -> str:
        content = match.group(1)
        return f"{content}）" if _should_strip_short_trailing_punctuation(content) else match.group(0)

    return re.sub(r"([^（）]*?)[，；：。！？、]+）", replace, text)


def _strip_short_trailing_punctuation(text: str) -> str:
    match = TRAILING_SEPARATOR_RE.search(text)
    if not match:
        return text
    return text[: match.start()] if _should_strip_short_trailing_punctuation(text[: match.start()]) else text


def _cleanup_punctuation_artifacts(text: str) -> str:
    current = PUNCTUATION_RUN_RE.sub(_collapse_punctuation_run, text)
    current = BRACKET_GROUP_SEPARATOR_RE.sub("）（", current)
    current = re.sub(r"（[，；：。！？、]+", "（", current)
    current = _strip_short_punctuation_before_closing_bracket(current)
    current = LEADING_PUNCTUATION_RE.sub("", current)
    current = _strip_short_trailing_punctuation(current)
    return current.strip()


def _outer_bracket_groups(text: str) -> list[tuple[int, int, str]]:
    groups: list[tuple[int, int, str]] = []
    depth = 0
    start: int | None = None

    for index, char in enumerate(text):
        if char == "（":
            if depth == 0:
                start = index
            depth += 1
            continue

        if char != "）" or depth == 0:
            continue

        depth -= 1
        if depth == 0 and start is not None:
            groups.append((start, index + 1, text[start + 1 : index]))
            start = None

    return groups


def _remove_delimited_duplicates(text: str) -> tuple[str, list[str]]:
    parts = DELIMITED_SPLIT_RE.split(text)
    seen: set[str] = set()
    duplicate_labels: list[str] = []
    rebuilt: list[str] = []

    for index in range(0, len(parts), 2):
        segment = parts[index] if index < len(parts) else ""
        delimiter = parts[index + 1] if index + 1 < len(parts) else ""
        trimmed = segment.strip()
        key = _duplicate_key(trimmed)

        if not key:
            if not rebuilt and delimiter:
                rebuilt.append(delimiter)
            continue
        if key in seen:
            duplicate_labels.append(trimmed)
            continue

        seen.add(key)
        rebuilt.append(f"{trimmed}{delimiter}")

    return "".join(rebuilt) or text, _unique(duplicate_labels)


def _remove_continuous_duplicates(text: str) -> tuple[str, list[str]]:
    current = text
    phrases: list[str] = []
    start = 0

    while start < len(current):
        matched = False
        max_length = min(12, (len(current) - start) // 2)
        for length in range(max_length, 1, -1):
            phrase = current[start : start + length]
            repeated = current[start + length : start + length * 2]
            if (
                phrase != repeated
                or phrase.isdigit()
                or not PHRASE_CHAR_RE.fullmatch(phrase)
            ):
                continue
            current = current[: start + length] + current[start + length * 2 :]
            phrases.append(phrase)
            matched = True
            # 删除后只需回看最大短语长度，避免每次都从字符串开头重新扫描。
            start = max(0, start - 12)
            break
        if not matched:
            start += 1

    return current, _unique(phrases)


def _check_duplicate_bracket_contents(text: str) -> tuple[str, list[str]]:
    seen: dict[str, str] = {}
    duplicate_contents: list[str] = []
    duplicate_ranges: list[tuple[int, int]] = []
    previous_group: tuple[int, int, str] | None = None

    for group in _outer_bracket_groups(text):
        start, end, content = group
        trimmed = content.strip()
        key = _duplicate_key(trimmed)
        if not key:
            continue
        if key in seen:
            duplicate_contents.append(seen[key])
            if previous_group is not None:
                previous_start, previous_end, previous_content = previous_group
                between = text[previous_end:start]
                if (
                    _duplicate_key(previous_content) == key
                    and not between.strip()
                ):
                    duplicate_ranges.append((start, end))
        else:
            seen[key] = trimmed
        previous_group = group

    if not duplicate_ranges:
        return text, _unique(duplicate_contents)

    rebuilt: list[str] = []
    cursor = 0
    for start, end in duplicate_ranges:
        rebuilt.append(text[cursor:start])
        cursor = end
    rebuilt.append(text[cursor:])
    current = "".join(rebuilt)
    return current, _unique(duplicate_contents)


def check_duplicates(text: str) -> tuple[str, list[str], bool]:
    bracket_text, bracket_contents = _check_duplicate_bracket_contents(text)
    delimited_text, duplicates = _remove_delimited_duplicates(bracket_text)
    current, phrases = _remove_continuous_duplicates(delimited_text)
    issues = [f"重复括号内容：{item}" for item in bracket_contents]
    issues.extend(f"重复内容：{item}" for item in duplicates)
    issues.extend(f"疑似连续重复：{item}" for item in phrases)
    return current, issues, current != text


def _brackets_are_unbalanced(text: str) -> bool:
    depth = 0
    for char in text:
        if char == "（":
            depth += 1
        elif char == "）":
            if depth == 0:
                return True
            depth -= 1
    return depth != 0


def check_bracket_issues(text: str) -> tuple[str, list[str], bool]:
    current = text
    issues: list[str] = []
    changed = False

    if EMPTY_BRACKET_RE.search(current):
        current = EMPTY_BRACKET_RE.sub("", current)
        issues.append("存在空括号")
        changed = True

    if REDUNDANT_NESTED_BRACKET_RE.search(current):
        while REDUNDANT_NESTED_BRACKET_RE.search(current):
            current = REDUNDANT_NESTED_BRACKET_RE.sub(
                lambda match: f"（{match.group(1).strip()}）", current
            )
        issues.append("存在嵌套括号：已去除重复外层括号")
        changed = True

    for wrong, correct in sorted(
        BRACKET_CONTENT_TYPO_MAP.items(), key=lambda item: len(item[0]), reverse=True
    ):
        pattern = re.compile(rf"（\s*{re.escape(wrong)}\s*）")
        if pattern.search(current):
            current = pattern.sub(f"（{correct}）", current)
            issues.append(f"括号内容疑似错字：{wrong} → {correct}")
            changed = True

    short_contents: list[str] = []
    for _, _, content in _outer_bracket_groups(current):
        content = re.sub(r"[\s\u3000]+", "", content)
        if 0 < len(content) <= 1:
            short_contents.append(content)

    if short_contents:
        issues.append(f"括号内容过短：{'、'.join(_unique(short_contents))}，请人工检查")

    if changed:
        current = _cleanup_punctuation_artifacts(current)

    return current, issues, changed


def check_physical_constraints(text: str) -> tuple[str, list[str], bool]:
    issues: list[str] = []

    for match in HEIGHT_WRONG_UNIT_RE.finditer(text):
        issues.append(f"身高单位疑似错误：{match.group(1)}{match.group(2)}")

    for match in HEIGHT_VALUE_RE.finditer(text):
        raw_value, unit = match.group(1), match.group(2)
        value = float(raw_value)
        height_cm = value * 100 if unit.lower() == "m" or unit == "米" else value
        if height_cm < 100 or height_cm > 230:
            issues.append(f"身高数值疑似异常：{raw_value}{unit}")

    for match in WEIGHT_WRONG_UNIT_RE.finditer(text):
        issues.append(f"体重单位疑似错误：{match.group(1)}{match.group(2)}")

    for match in WEIGHT_VALUE_RE.finditer(text):
        raw_value, unit = match.group(1), match.group(2)
        value = float(raw_value)
        weight_kg = value / 2 if unit == "斤" else value
        if weight_kg < 30 or weight_kg > 200:
            issues.append(f"体重数值疑似异常：{raw_value}{unit}")

    return text, _unique(issues), False


def check_tuition(text: str) -> tuple[str, list[str], bool]:
    issues: list[str] = []

    for match in TUITION_RE.finditer(text):
        raw_value, unit = match.group(1), match.group(2)
        value = float(raw_value)
        yuan = value * 10000 if unit in {"万", "万元"} else value
        if yuan > 300000:
            issues.append(f"学费金额疑似异常：{raw_value}{unit}超过30万元")

    return text, _unique(issues), False


def check_format_issues(text: str) -> tuple[str, list[str], bool]:
    current = text
    issues: list[str] = []
    changed = False

    if "\n" in current or "\r" in current:
        current = re.sub(
            r"[\t \u00A0\u3000]*[\r\n]+[\t \u00A0\u3000]*", "；", current
        )
        # 原句已有结尾标点时，换行只表示视觉分隔，不叠加分号。
        current = re.sub(r"([，；：。！？、])；", r"\1", current)
        current = re.sub(r"；([，；：。！？、])", r"\1", current)
        issues.append("换行符已统一为中文分号")
        changed = True

    if HORIZONTAL_SPACE_RE.search(current):
        before = current
        current = HORIZONTAL_SPACE_RE.sub(" ", current.strip())
        current = HAN_SPACE_HAN_RE.sub("", current)
        current = re.sub(r"\s*([，；：。！？、（）])\s*", r"\1", current)
        if current != before:
            issues.append("存在多余空格")
            changed = True
    else:
        trimmed = current.strip()
        if trimmed != current:
            current = trimmed
            issues.append("存在多余空格")
            changed = True

    if re.search(r"[,;:()]", current):
        current = current.translate(ENGLISH_PUNCTUATION_MAP)
        issues.append("英文标点已统一为中文标点")
        changed = True

    if PUNCTUATION_RUN_RE.search(current):
        current = PUNCTUATION_RUN_RE.sub(_collapse_punctuation_run, current)
        issues.append("存在连续标点")
        changed = True

    before_bracket_separator = current
    current = BRACKET_GROUP_SEPARATOR_RE.sub("）（", current)
    if current != before_bracket_separator:
        issues.append("括号组之间存在多余标点符号")
        changed = True

    before_extra_punctuation = current
    current = re.sub(r"（[，；：。！？、]+", "（", current)
    current = _strip_short_punctuation_before_closing_bracket(current)
    current = LEADING_PUNCTUATION_RE.sub("", current)
    current = _strip_short_trailing_punctuation(current)
    if current != before_extra_punctuation:
        issues.append("存在多余标点符号")
        changed = True

    if _brackets_are_unbalanced(current):
        issues.append("括号疑似不成对，请人工检查")

    return current, issues, changed


def normalize_remark(text: Any) -> str:
    """执行安全格式标准化，不应用错字和重复内容规则。"""
    if _is_empty_remark(text):
        return ""
    current, _, _ = check_abnormal_chars(_to_text(text))
    current, _, _ = check_format_issues(current)
    return current


def process_remark(text: Any) -> dict[str, str]:
    """检查单条备注，返回问题标注和仅在发生自动修正时填写的备注。"""
    if _is_empty_remark(text) or _is_invalid_remark(text):
        return {"issues": "", "fixed": ""}

    original = _to_text(text)
    current, abnormal_issues, abnormal_changed = check_abnormal_chars(original)
    current, ocr_noise_issues, ocr_noise_changed = check_ocr_noise_symbols(current)
    current, typo_issues, typo_changed = check_typos(current)
    current, suspect_typo_issues, suspect_typo_changed = check_suspect_typos(current)
    current, format_issues, format_changed = check_format_issues(current)
    current, bracket_issues, bracket_changed = check_bracket_issues(current)
    current, duplicate_issues, duplicate_changed = check_duplicates(current)
    current, physical_issues, physical_changed = check_physical_constraints(current)
    current, tuition_issues, tuition_changed = check_tuition(current)

    issue_list = _unique(
        typo_issues
        + ocr_noise_issues
        + suspect_typo_issues
        + format_issues
        + bracket_issues
        + duplicate_issues
        + physical_issues
        + tuition_issues
        + abnormal_issues
    )
    auto_changed = (
        abnormal_changed
        or ocr_noise_changed
        or typo_changed
        or suspect_typo_changed
        or format_changed
        or bracket_changed
        or duplicate_changed
        or physical_changed
        or tuition_changed
    )
    fixed = current if auto_changed and current != original else ""
    return {"issues": "；".join(issue_list), "fixed": fixed}


def _copy_header_style(source_cell: Any, target_cell: Any) -> None:
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.font:
        target_cell.font = copy(source_cell.font)
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)


def apply_uniform_excel_export_formatting(workbook: Any) -> None:
    """统一导出列宽与水平对齐，并保留单元格已有的其他对齐属性。"""
    for worksheet in workbook.worksheets:
        for column_index in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(column_index)
            worksheet.column_dimensions[column_letter].width = EXCEL_EXPORT_COLUMN_WIDTH

        # 仅遍历已实例化单元格，避免稀疏工作表因整块遍历而急剧膨胀内存和文件体积。
        for cell in list(worksheet._cells.values()):
            alignment = copy(cell.alignment)
            alignment.horizontal = "left"
            if alignment.vertical is None:
                alignment.vertical = "center"
            cell.alignment = alignment


def process_excel(input_path: str | Path) -> Path:
    """检查工作簿中所有包含标准备注列的工作表，并输出新文件。"""
    path = Path(input_path).expanduser().resolve()
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"输入文件不存在：{path}")
    if path.suffix.lower() != ".xlsx":
        raise ValueError("仅支持 .xlsx 文件")

    output_path = path.with_name(f"{path.stem}_备注检查结果.xlsx")
    output_index = 1
    while output_path.exists():
        output_path = path.with_name(f"{path.stem}_备注检查结果_{output_index}.xlsx")
        output_index += 1
    workbook = load_workbook(path)
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    total_rows = 0
    remark_rows = 0
    problem_rows = 0
    fixed_rows = 0
    processed_sheets = 0

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        existing_cells = list(worksheet._cells.values())
        headers = [cell.value for cell in worksheet[1]]
        try:
            remark_column = find_remark_column(headers)
        except ValueError:
            continue

        processed_sheets += 1
        last_value_row = max(
            (cell.row for cell in existing_cells if cell.value is not None),
            default=1,
        )
        data_row_count = max(0, last_value_row - 1)
        total_rows += data_row_count

        remark_col_index = next(
            index
            for index, value in enumerate(headers, start=1)
            if str(value).strip() == remark_column
        )
        last_data_col = max(worksheet.max_column, remark_col_index)
        issue_col_index = last_data_col + 1
        fixed_col_index = last_data_col + 2

        source_header = worksheet.cell(row=1, column=remark_col_index)
        issue_header = worksheet.cell(row=1, column=issue_col_index, value="备注问题标注")
        fixed_header = worksheet.cell(row=1, column=fixed_col_index, value="修改后备注")
        _copy_header_style(source_header, issue_header)
        _copy_header_style(source_header, fixed_header)

        remark_source_cells = sorted(
            (
                cell
                for cell in existing_cells
                if cell.row >= 2
                and cell.column == remark_col_index
                and cell.value is not None
            ),
            key=lambda cell: cell.row,
        )

        for source_cell in remark_source_cells:
            row_offset = source_cell.row
            value = source_cell.value
            result = process_remark(value)
            issue_cell = worksheet.cell(row=row_offset, column=issue_col_index, value=result["issues"])
            fixed_cell = worksheet.cell(row=row_offset, column=fixed_col_index, value=result["fixed"])
            issue_cell.alignment = Alignment(wrap_text=True, vertical="top")
            fixed_cell.alignment = Alignment(wrap_text=True, vertical="top")

            if not _is_empty_remark(value):
                remark_rows += 1
            if result["issues"]:
                problem_rows += 1
                issue_cell.fill = yellow_fill
                fixed_cell.fill = yellow_fill
            if result["fixed"]:
                fixed_rows += 1

        worksheet.column_dimensions[get_column_letter(issue_col_index)].width = 48
        worksheet.column_dimensions[get_column_letter(fixed_col_index)].width = 42

    if processed_sheets == 0:
        expected = "、".join(REMARK_COLUMN_NAMES)
        raise ValueError(f"所有工作表中均未找到备注列，支持的列名：{expected}")

    apply_uniform_excel_export_formatting(workbook)
    workbook.save(output_path)
    print(f"总行数：{total_rows}")
    print(f"有备注行数：{remark_rows}")
    print(f"检测出问题的行数：{problem_rows}")
    print(f"自动生成修改后备注的行数：{fixed_rows}")
    print(f"输出文件路径：{output_path}")
    return output_path


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="检查招生计划 Excel 备注字段")
    parser.add_argument("input_path", nargs="?", help="待检查的 .xlsx 文件路径")
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    input_path = args.input_path
    if not input_path:
        input_path = input("请输入 Excel 文件路径：").strip().strip('"')

    try:
        process_excel(input_path)
    except Exception as exc:  # 命令行入口需要给用户可读错误，而不是整段堆栈。
        print(f"处理失败：{exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
