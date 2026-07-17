from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from check_remarks import process_excel, process_remark


class ProcessRemarkTests(unittest.TestCase):
    def test_typo_is_labeled_and_fixed(self) -> None:
        result = process_remark("详见院校招生张程")
        self.assertIn("疑似错字：详见院校招生张程 → 详见院校招生章程", result["issues"])
        self.assertEqual(result["fixed"], "详见院校招生章程")

    def test_ocr_typos_are_labeled_and_fixed(self) -> None:
        result = process_remark("不招色育，项目选拔进人，详贝院校招生章程")
        self.assertIn("疑似错字：色育 → 色盲", result["issues"])
        self.assertIn("疑似错字：进人 → 进入", result["issues"])
        self.assertIn("疑似错字：详贝院校招生章程 → 详见院校招生章程", result["issues"])
        self.assertEqual(result["fixed"], "不招色盲，项目选拔进入，详见院校招生章程")

    def test_more_ocr_typos_are_labeled_and_fixed(self) -> None:
        result = process_remark(
            "只召英语语仲考牛，身体建康，成绩台格，国家专顷计划，学贵310000元"
        )
        self.assertIn("疑似错字：只召 → 只招", result["issues"])
        self.assertIn("疑似错字：语仲 → 语种", result["issues"])
        self.assertIn("疑似错字：考牛 → 考生", result["issues"])
        self.assertIn("疑似错字：身体建康 → 身体健康", result["issues"])
        self.assertIn("疑似错字：成绩台格 → 成绩合格", result["issues"])
        self.assertIn("疑似错字：国家专顷计划 → 国家专项计划", result["issues"])
        self.assertIn("疑似错字：学贵 → 学费", result["issues"])
        self.assertIn("学费金额疑似异常：310000元超过30万元", result["issues"])
        self.assertEqual(
            result["fixed"],
            "只招英语语种考生，身体健康，成绩合格，国家专项计划，学费310000元",
        )

    def test_low_confidence_ocr_typos_are_only_labeled(self) -> None:
        result = process_remark("只招英语老生，慎填")
        self.assertIn("疑似 OCR 错字：老生 可能为 考生，请人工检查", result["issues"])
        self.assertIn("疑似 OCR 错字：慎填 可能为 慎报，请人工检查", result["issues"])
        self.assertEqual(result["fixed"], "")

    def test_delimited_duplicate_is_removed(self) -> None:
        result = process_remark("不招色盲；不招色盲；详见招生章程")
        self.assertIn("重复内容：不招色盲", result["issues"])
        self.assertEqual(result["fixed"], "不招色盲；详见招生章程")

    def test_sentence_duplicate_keeps_one_terminal_mark(self) -> None:
        result = process_remark("不招色盲。不招色盲。")
        self.assertEqual(result["fixed"], "不招色盲。")

    def test_continuous_duplicate_is_removed(self) -> None:
        result = process_remark("详见详见招生章程")
        self.assertIn("疑似连续重复：详见", result["issues"])
        self.assertEqual(result["fixed"], "详见招生章程")

    def test_adjacent_duplicate_bracket_content_is_labeled_and_removed(self) -> None:
        result = process_remark("（师范类）（师范类）")
        self.assertEqual(result["issues"], "重复括号内容：师范类")
        self.assertEqual(result["fixed"], "（师范类）")

    def test_non_adjacent_duplicate_bracket_content_is_only_labeled(self) -> None:
        result = process_remark("（海曙校区）（第一学年）（海曙校区）")
        self.assertEqual(result["issues"], "重复括号内容：海曙校区")
        self.assertEqual(result["fixed"], "")

    def test_empty_nested_and_short_brackets_are_checked(self) -> None:
        empty = process_remark("（）详见招生章程")
        self.assertEqual(empty["issues"], "存在空括号")
        self.assertEqual(empty["fixed"], "详见招生章程")

        nested = process_remark("（（师范类））")
        self.assertEqual(nested["issues"], "存在嵌套括号：已去除重复外层括号")
        self.assertEqual(nested["fixed"], "（师范类）")

        short = process_remark("（男）只招")
        self.assertEqual(short["issues"], "括号内容过短：男，请人工检查")
        self.assertEqual(short["fixed"], "")

    def test_major_list_parentheses_are_not_marked_as_duplicate_or_nested(self) -> None:
        normal = process_remark(
            "（只承认教育部规定的全国性加分政策）（包含专业：教育学、融合教育（非公费师范）、特殊教育（非公费师范））"
        )
        self.assertEqual(normal["issues"], "")
        self.assertEqual(normal["fixed"], "")

        ai = process_remark(
            "（管理与经济方向）（认同并执行四川省少数民族地区加分项目和分值）（包含专业：工商管理（含数字创新管理方向）、工商管理（含AI双学位）、经济学、国际经济与贸易、经济学（含AI双学位））"
        )
        self.assertEqual(ai["issues"], "")
        self.assertEqual(ai["fixed"], "")

    def test_height_and_weight_abnormal_values_are_only_labeled(self) -> None:
        result = process_remark("身高不低于80cm，体重不超过500kg")
        self.assertIn("身高数值疑似异常：80cm", result["issues"])
        self.assertIn("体重数值疑似异常：500kg", result["issues"])
        self.assertEqual(result["fixed"], "")

        unit = process_remark("身高不低于170kg，体重不超过170cm")
        self.assertIn("身高单位疑似错误：170kg", unit["issues"])
        self.assertIn("体重单位疑似错误：170cm", unit["issues"])
        self.assertEqual(unit["fixed"], "")

    def test_tuition_abnormal_values_are_only_labeled(self) -> None:
        normal = process_remark("学费300000元")
        self.assertEqual(normal["issues"], "")
        self.assertEqual(normal["fixed"], "")

        yuan = process_remark("学费300001元")
        self.assertEqual(yuan["issues"], "学费金额疑似异常：300001元超过30万元")
        self.assertEqual(yuan["fixed"], "")

        ten_thousand = process_remark("学费31万元")
        self.assertEqual(ten_thousand["issues"], "学费金额疑似异常：31万元超过30万元")
        self.assertEqual(ten_thousand["fixed"], "")

    def test_ocr_noise_symbol_and_bracket_group_separator_are_fixed(self) -> None:
        result = process_remark(
            "（按智慧安全方向培%养，成龙校区）、（北斗应用与低空经济现代产业学院、九江）（前2年在南昌校区就读，后2年在九江校区就读）"
        )
        self.assertIn("疑似 OCR 多余符号：%/％", result["issues"])
        self.assertIn("括号组之间存在多余标点符号", result["issues"])
        self.assertEqual(
            result["fixed"],
            "（按智慧安全方向培养，成龙校区）（北斗应用与低空经济现代产业学院、九江）（前2年在南昌校区就读，后2年在九江校区就读）",
        )

    def test_unusual_bracket_content_typo_is_fixed(self) -> None:
        result = process_remark("（不招色盲、色弱考生）（通类）")
        self.assertEqual(result["issues"], "括号内容疑似错字：通类 → 普通类")
        self.assertEqual(result["fixed"], "（不招色盲、色弱考生）（普通类）")

    def test_format_cleanup_preserves_meaning(self) -> None:
        result = process_remark("不招 色盲;详见招生章程。。\n不招色弱；；")
        self.assertIn("存在多余空格", result["issues"])
        self.assertIn("英文标点已统一为中文标点", result["issues"])
        self.assertIn("存在连续标点", result["issues"])
        self.assertIn("存在多余标点符号", result["issues"])
        self.assertEqual(result["fixed"], "不招色盲；详见招生章程。不招色弱")

    def test_long_sentence_trailing_punctuation_is_preserved(self) -> None:
        result = process_remark("录取规则详见学校招生章程；")
        self.assertNotIn("存在多余标点符号", result["issues"])
        self.assertEqual(result["fixed"], "")

        bracket = process_remark("（具体要求详见学校招生章程。）")
        self.assertNotIn("存在多余标点符号", bracket["issues"])
        self.assertEqual(bracket["fixed"], "")

    def test_short_phrase_trailing_punctuation_is_flagged(self) -> None:
        result = process_remark("师范类。")
        self.assertIn("存在多余标点符号", result["issues"])
        self.assertEqual(result["fixed"], "师范类")

        bracket = process_remark("（师范类。）")
        self.assertIn("存在多余标点符号", bracket["issues"])
        self.assertEqual(bracket["fixed"], "（师范类）")

    def test_uncertain_issues_are_not_force_fixed(self) -> None:
        bracket = process_remark("只招英语考生（口试成绩合格")
        self.assertEqual(bracket["issues"], "括号疑似不成对，请人工检查")
        self.assertEqual(bracket["fixed"], "")

        abnormal = process_remark("学费□元")
        self.assertEqual(abnormal["issues"], "存在疑似异常字符：□")
        self.assertEqual(abnormal["fixed"], "")

    def test_empty_and_invalid_values_are_ignored(self) -> None:
        for value in (None, "", "nan", "None", "无", "暂无", "-", "/"):
            self.assertEqual(process_remark(value), {"issues": "", "fixed": ""})


class ProcessExcelTests(unittest.TestCase):
    def test_missing_file_and_wrong_extension_are_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            missing_path = Path(temp_dir) / "missing.xlsx"
            wrong_extension = Path(temp_dir) / "plan.xls"
            wrong_extension.write_bytes(b"not-an-xlsx")

            with self.assertRaisesRegex(FileNotFoundError, "输入文件不存在"):
                process_excel(missing_path)
            with self.assertRaisesRegex(ValueError, "仅支持 .xlsx 文件"):
                process_excel(wrong_extension)

    def test_workbook_without_remark_column_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "无备注列.xlsx"
            workbook = Workbook()
            workbook.active.append(["专业", "学校"])
            workbook.active.append(["计算机", "测试大学"])
            workbook.save(input_path)

            with self.assertRaisesRegex(ValueError, "所有工作表中均未找到备注列"):
                process_excel(input_path)

    def test_sparse_workbook_does_not_expand_empty_rows(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "稀疏计划.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["备注"])
            sheet["A100000"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
            workbook.save(input_path)

            output_path = process_excel(input_path)
            result_book = load_workbook(output_path)
            result_sheet = result_book.active

            self.assertLess(len(result_sheet._cells), 20)
            self.assertEqual(result_sheet["B1"].value, "备注问题标注")
            self.assertEqual(result_sheet["C1"].value, "修改后备注")

    def test_existing_output_file_is_not_overwritten(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "招生计划.xlsx"
            existing_output = Path(temp_dir) / "招生计划_备注检查结果.xlsx"

            workbook = Workbook()
            workbook.active.append(["备注"])
            workbook.active.append(["只招英语语种考生"])
            workbook.save(input_path)
            existing_output.write_bytes(b"existing-result")

            output_path = process_excel(input_path)

            self.assertEqual(existing_output.read_bytes(), b"existing-result")
            self.assertEqual(output_path.name, "招生计划_备注检查结果_1.xlsx")
            self.assertTrue(output_path.exists())

    def test_workbook_is_copied_and_result_cells_are_highlighted(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "招生计划.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "计划"
            sheet.append(["专业", "备注"])
            sheet.append(["计算机", "不招色盲。不招色盲。"])
            sheet.append(["英语", "只招英语语种考生"])
            sheet["A1"].alignment = Alignment(
                horizontal="center",
                vertical="top",
                wrap_text=True,
                shrink_to_fit=True,
                indent=2,
            )
            description_sheet = workbook.create_sheet("说明")
            description_sheet["A1"] = "原工作表应保留"
            description_sheet["A1"].alignment = Alignment(horizontal="center")
            workbook.save(input_path)

            output_path = process_excel(input_path)
            self.assertTrue(input_path.exists())
            self.assertTrue(output_path.exists())
            self.assertNotEqual(input_path, output_path)

            result_book = load_workbook(output_path)
            self.assertEqual(result_book.sheetnames, ["计划", "说明"])
            result_sheet = result_book["计划"]
            self.assertEqual(result_sheet["C1"].value, "备注问题标注")
            self.assertEqual(result_sheet["D1"].value, "修改后备注")
            self.assertEqual(result_sheet["D2"].value, "不招色盲。")
            self.assertEqual(result_sheet["C2"].fill.fgColor.rgb, "00FFF2CC")
            self.assertIsNone(result_sheet["C3"].value)
            self.assertIsNone(result_sheet["D3"].value)

            for result_sheet in result_book.worksheets:
                for column_index in range(1, result_sheet.max_column + 1):
                    column_letter = get_column_letter(column_index)
                    self.assertEqual(
                        result_sheet.column_dimensions[column_letter].width,
                        12.75,
                    )

                for row in result_sheet.iter_rows():
                    for cell in row:
                        self.assertEqual(cell.alignment.horizontal, "left")
                        self.assertIsNotNone(cell.alignment.vertical)

            preserved_alignment = result_book["计划"]["A1"].alignment
            self.assertEqual(preserved_alignment.vertical, "top")
            self.assertTrue(preserved_alignment.wrap_text)
            self.assertTrue(preserved_alignment.shrink_to_fit)
            self.assertEqual(preserved_alignment.indent, 2)
            self.assertEqual(result_book["说明"]["A1"].alignment.vertical, "center")


if __name__ == "__main__":
    unittest.main()
